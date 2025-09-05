#!/usr/bin/env python3
# eclass_graph_loader.py
"""
Convert an Excel workbook of sensor specifications into a Neo4j graph governed by
the ECLASS Advanced taxonomy (v15.0).

Runtime & style:
- Python 3.11
- Type hints, structured logging (JSON), clear exceptions
- Idempotent: safe to re-run (MERGE-based upserts; stable IDs)
- Efficient on large sheets (openpyxl streaming chunk reader; batched Cypher)
- Dependencies (from existing requirements only): pandas, openpyxl, python-dotenv,
  neo4j, regex, rapidfuzz, pint, tqdm, pyyaml, orjson

CLI:
  --excel PATH (required unless --dry-run with fixture)
  --sheet NAME (default="ПКШ-ОБОР")
  --idcol NAME (default="Идентификатор")
  --config PATH (YAML/JSON with column mapping & RU/EN keyword dictionaries)
  --eclass-classes PATH (CSV/JSON of classes)
  --eclass-props PATH (CSV/JSON of class->properties dictionary)
  --batch-size INT (default=1000)
  --lang LIST (default="ru,en")
  --confidence-threshold FLOAT (default=0.75)
  --dry-run (write CSV/Parquet only; no DB writes)
  --check-db (connect, ensure constraints, report OK, exit)

Graph model (Neo4j):
  Nodes:
    ECLASSClass {code, label, version}
    Property {irdi, name, version, datatype, canonicalUnit}
    Unit {symbol, name}
    Item {lineId, role, specCore, qty, uom, flags (labels :Sensor/:Assembly)}
    PropertyValue {pvId, numericValue|textValue|boolValue|enumValue, unit, source, confidence}
    SourceRow {rowId, sheet, excelPath, rawText, hash}
  Relationships:
    (Item)-[:INSTANCE_OF]->(ECLASSClass)
    (ECLASSClass)-[:HAS_PROPERTY_MODEL]->(Property)
    (Item)-[:HAS_PROPERTY]->(PropertyValue)-[:OF_PROPERTY]->(Property)
    (PropertyValue)-[:USES_UNIT]->(Unit)
    (Item)-[:DERIVED_FROM]->(SourceRow)
    (Parent:Item)-[:HAS_COMPONENT {qty}]->(Component:Item)
    (Accessory:Item)-[:ACCESSORY_OF]->(Main:Item)
    (Service:Item)-[:SERVICE_FOR]->(Main:Item)
    (Item)-[:REQUIRES_REVIEW]->(SourceRow) when low confidence

Constraints (Neo4j 5+):
  ECLASSClass(code, version) UNIQUE
  Property(irdi, version) UNIQUE
  Unit(symbol) UNIQUE
  Item(lineId) UNIQUE
  SourceRow(hash) UNIQUE
"""

from __future__ import annotations

import argparse
import hashlib
import itertools
import math
import os
import sys
import time
import re
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Callable, Dict, Generator, Iterable, List, Optional, Tuple

# --- Dependency checks first (only allowed libraries) ------------------------------------------
def check_deps() -> None:
    """
    Ensure all required third-party dependencies are available.
    Raises ImportError with a helpful message listing missing modules.
    """
    required = [
        "pandas",
        "openpyxl",
        "dotenv",
        "neo4j",
        "regex",
        "rapidfuzz",
        "pint",
        "tqdm",
        "yaml",   # pyyaml
        "orjson",
    ]
    missing: List[str] = []
    for mod in required:
        try:
            __import__(mod)
        except Exception:
            missing.append(mod)
    if missing:
        raise ImportError(
            "Missing required dependencies: "
            + ", ".join(missing)
            + ". Please install them as per your project's requirements.txt."
        )

check_deps()  # Fail fast

# Now import the verified libs
import orjson
import pandas as pd
from dotenv import load_dotenv
from neo4j import GraphDatabase, Driver, Session
import regex as re
from rapidfuzz import fuzz, process as rf_process
from pint import UnitRegistry
from tqdm import tqdm
import yaml  # type: ignore

# Add the directory containing utils.py to the Python path
utils_dir = os.path.expanduser("~/Projects/Sensors-kb")  # Path to DIRECTORY containing utils.py
sys.path.append(utils_dir)

# Local utils loader with type hints for Pylance
try:
    from utils import load_model  # type: ignore
except ImportError:
    load_model = None  # type: ignore
    _UTILS_IMPORT_ERROR = "utils module not found"
else:
    _UTILS_IMPORT_ERROR = None

# --- Structured JSON logging -------------------------------------------------------------------
import logging

class JsonFormatter(logging.Formatter):
    def format(self, record: logging.LogRecord) -> str:
        payload = {
            "ts": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime(record.created)),
            "level": record.levelname,
            "name": record.name,
            "msg": record.getMessage(),
        }
        if record.exc_info:
            payload["exc_info"] = self.formatException(record.exc_info)
        if hasattr(record, "extra"):
            try:
                payload["extra"] = record.extra  # type: ignore[attr-defined]
            except Exception:
                pass
        return orjson.dumps(payload).decode("utf-8")

def configure_logging(verbosity: int = 0) -> None:
    level = logging.DEBUG if verbosity >= 2 else logging.INFO if verbosity == 1 else logging.INFO
    handler = logging.StreamHandler(sys.stdout)
    handler.setFormatter(JsonFormatter())
    root = logging.getLogger()
    root.handlers.clear()
    root.addHandler(handler)
    root.setLevel(level)

logger = logging.getLogger("eclass_loader")

# --- Data classes ------------------------------------------------------------------------------

@dataclass
class ItemCandidate:
    lineId: str
    groupId: str
    parentLineId: Optional[str]
    role: str  # parent|component|accessory|service
    qty: float
    uom: Optional[str]
    specCore: str
    rawText: str
    rowHash: str
    rowIndex: int
    sheet: str
    excelPath: str
    flags: List[str]

@dataclass
class PropertyValueCandidate:
    irdi: Optional[str]
    name: str
    value: Any
    unit: Optional[str]
    datatype: str  # number|string|boolean|enum
    source: str    # "llm" or "regex"
    confidence: float
    canonicalUnit: Optional[str] = None

@dataclass
class ClassifiedItem:
    lineId: str
    eclass_code: Optional[str]
    eclass_label: Optional[str]
    confidence: float
    properties: List[PropertyValueCandidate]
    notes: str

# --- Helpers ------------------------------------------------------------------------------------

def sha1_hex(text: str) -> str:
    return hashlib.sha1(text.encode("utf-8")).hexdigest()

def env(key: str) -> str:
    v = os.getenv(key)
    if not v:
        raise RuntimeError(f"Missing required environment variable: {key}")
    return v

def safe_float(x: Any, default: float = 1.0) -> float:
    try:
        if isinstance(x, (int, float)):
            return float(x)
        s = str(x).strip().replace(",", ".")
        return float(s)
    except Exception:
        return default

def to_list_langs(s: str | Iterable[str]) -> List[str]:
    if isinstance(s, str):
        return [x.strip() for x in s.split(",") if x.strip()]
    return [str(x).strip() for x in s if str(x).strip()]

def mask_secret(v: str) -> str:
    if not v:
        return v
    if len(v) <= 6:
        return "***"
    return v[:2] + "***" + v[-2:]

def orjson_loads(data: str | bytes) -> Any:
    if isinstance(data, str):
        return orjson.loads(data)
    return orjson.loads(data)

def strip(value: Any) -> Optional[str]:
    """Safely strip a value or return None."""
    if value is None:
        return None
    if not isinstance(value, str):
        value = str(value)
    return value.strip()

def fuzzy_get(row: pd.Series, colname: str) -> Optional[str]:
    if (colname := strip(colname)) in row.index:
        return strip(row.get(colname))
    target = re.sub(r'\s+', '', colname.lower())
    for c in row.index:
        if re.sub(r'\s+', '', str(c).lower()) == target:
            return strip(row.get(c))  # Added missing return statement
    return None

def chunked(seq, size):
    """Split a sequence into chunks of specified size."""
    return (seq[i:i+size] for i in range(0, len(seq), size))

# --- Config & dictionaries ----------------------------------------------------------------------

def load_config(path: Optional[Path]) -> Dict[str, Any]:
    """
    Load optional YAML/JSON config: column mappings, RU/EN keyword dictionaries, unit aliases.
    """
    default = {
        "columns": {
            "id": "Идентификатор",
            # The core text/spec column will be guessed if not specified:
            # tries 'Описание', 'Описание изделия', 'Description', 'Spec', 'Specification'
            "text_candidates": ["Описание", "Описание изделия", "Description", "Spec", "Specification"],
        },
        "keywords": {
            "assembly": ["в сборе", "комплект", "узел", "в комплекте", "assembly", "kit", "set"],
            "accessory": ["аксессуар", "accessory"],
            "service": ["услуга", "service"],
            "component_markers": ["с", "включая", "в составе", "including", "with"],  # may split child parts
            "sensor_like": ["датчик", "сенсор", "преобразователь", "transmitter", "sensor", "probe"],
        },
        "unit_aliases": {
            "°С": "degC", "°C": "degC", "град.С": "degC",
            "мА": "mA", "ма": "mA", "A": "A", "В": "V", "в": "V",
            "бар": "bar", "мбар": "mbar", "Па": "Pa", "кПа": "kPa", "МПа": "MPa",
            "шт": "count", "pcs": "count",
        },
        "eclass_version": "15.0",
    }
    if not path:
        return default
    p = Path(path)
    if not p.exists():
        return default
    try:
        if p.suffix.lower() in {".yaml", ".yml"}:
            with p.open("rt", encoding="utf-8") as fh:
                cfg = yaml.safe_load(fh)
        else:
            cfg = orjson.loads(p.read_bytes())
        if not isinstance(cfg, dict):
            return default
        # merge shallowly
        merged = default.copy()
        merged.update(cfg)
        if "columns" in cfg:
            merged["columns"].update(cfg["columns"])
        if "keywords" in cfg:
            merged["keywords"].update(cfg["keywords"])
        if "unit_aliases" in cfg:
            merged["unit_aliases"].update(cfg["unit_aliases"])
        return merged
    except Exception as e:
        logger.warning("Failed to load config; using defaults", extra={"extra": {"error": str(e)}})
        return default

def load_eclass_classes(path: Optional[Path]) -> Dict[str, Dict[str, Any]]:
    """
    Load ECLASS classes into dict by 8-digit code "XX-XX-XX-XX".
    Supported formats: JSON[{code,label,version,parent}], CSV with same columns.
    """
    classes: Dict[str, Dict[str, Any]] = {}
    if not path:
        return classes
    p = Path(path)
    if not p.exists():
        return classes
    try:
        if p.suffix.lower() == ".json":
            data = orjson.loads(p.read_bytes())
            if isinstance(data, dict) and "classes" in data:
                data = data["classes"]
        else:
            data = pd.read_csv(p).to_dict(orient="records")
        for row in data:
            code = str(row.get("code") or "").strip()
            if not code:
                continue
            classes[code] = {
                "code": code,
                "label": (row.get("label") or "").strip(),
                "version": str(row.get("version") or "15.0"),
                "parent": (row.get("parent") or "").strip() or None,
            }
        return classes
    except Exception as e:
        logger.error("Failed to load ECLASS classes", extra={"extra": {"path": str(p), "error": str(e)}})
        return {}

def load_eclass_properties(path: Optional[Path]) -> Dict[str, List[Dict[str, Any]]]:
    """
    Load ECLASS property dictionary.
    Returns dict[class_code] = list of {irdi, name, datatype, canonicalUnit, allowedValues, version}
    Formats:
      - JSON { "byClass": { "XX-XX-XX-XX": [ ... ] } } or list with class_code field
      - CSV with columns: class_code, irdi, name, datatype, canonicalUnit, allowedValues, version
    """
    if not path:
        return {}
    p = Path(path)
    if not p.exists():
        return {}
    try:
        by_class: Dict[str, List[Dict[str, Any]]] = {}
        if p.suffix.lower() == ".json":
            data = orjson.loads(p.read_bytes())
            if isinstance(data, dict) and "byClass" in data:
                by_class = data["byClass"]
            elif isinstance(data, list):
                for row in data:
                    c = str(row.get("class_code") or "").strip()
                    if not c:
                        continue
                    by_class.setdefault(c, []).append({
                        "irdi": (row.get("irdi") or "").strip(),
                        "name": (row.get("name") or "").strip(),
                        "datatype": (row.get("datatype") or "string").strip(),
                        "canonicalUnit": (row.get("canonicalUnit") or None),
                        "allowedValues": row.get("allowedValues") or None,
                        "version": str(row.get("version") or "15.0"),
                    })
            else:
                by_class = {}
        else:
            df = pd.read_csv(p)
            for _, row in df.iterrows():
                c = str(row.get("class_code") or "").strip()
                if not c:
                    continue
                by_class.setdefault(c, []).append({
                    "irdi": str(row.get("irdi") or "").strip(),
                    "name": str(row.get("name") or "").strip(),
                    "datatype": str(row.get("datatype") or "string").strip(),
                    "canonicalUnit": (row.get("canonicalUnit") if pd.notna(row.get("canonicalUnit")) else None),
                    "allowedValues": (row.get("allowedValues") if pd.notna(row.get("allowedValues")) else None),
                    "version": str(row.get("version") or "15.0").strip(),
                })
        # Normalize keys and lists
        out: Dict[str, List[Dict[str, Any]]] = {}
        for c, lst in by_class.items():
            if not isinstance(lst, list):
                continue
            out[str(c).strip()] = lst
        return out
    except Exception as e:
        logger.error("Failed to load ECLASS properties", extra={"extra": {"path": str(p), "error": str(e)}})
        return {}

# --- Excel streaming reader ---------------------------------------------------------------------

def read_excel_chunks(path: Path, sheet_name: str, chunk_size: int = 1000) -> Generator[pd.DataFrame, None, None]:
    """
    Stream rows from an Excel sheet in chunks using openpyxl read_only mode.
    Returns DataFrames with header from the first row.
    """
    from openpyxl import load_workbook  # type: ignore

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{sheet_name}' not found in {path.name}. Available: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return
    header = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(header)]
    batch: List[List[Any]] = []
    row_index = 1  # considering first row is header
    for r in rows_iter:
        row_index += 1
        batch.append(list(r))
        if len(batch) >= chunk_size:
            df = pd.DataFrame(batch, columns=header)
            df["__row_index__"] = list(range(row_index - len(batch) + 1, row_index + 1))
            yield df
            batch.clear()
    if batch:
        df = pd.DataFrame(batch, columns=header)
        df["__row_index__"] = list(range(row_index - len(batch) + 1, row_index + 1))
        yield df

def guess_text_column(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    cols_lower = {c.lower(): c for c in df.columns}
    for cand in candidates:
        if cand.lower() in cols_lower:
            return cols_lower[cand.lower()]
    # Heuristic: prefer any column containing 'опис' or 'desc' or 'spec'
    for c in df.columns:
        cl = str(c).lower()
        if "опис" in cl or "desc" in cl or "spec" in cl:
            return c
    # Fallback to the widest text-like column
    return None

# --- Row splitting & heuristics -----------------------------------------------------------------

RANGE_RE = re.compile(
    r"(?P<min>-?\d+(?:[.,]\d+)?)\s*(?:–|-|—|to)\s*(?P<max>-?\d+(?:[.,]\d+)?)\s*(?P<unit>[°%A-Za-zµμΩOhm/·\s]*[A-Za-z%°Ω])?",
    re.IGNORECASE,
)
OUTPUT_SIG_RE = re.compile(
    r"(4\s*[–-]\s*20\s*mA|0\s*[–-]\s*10\s*V|HART|Profibus|Modbus|Foundation\s*Fieldbus|RS[- ]?485)",
    re.IGNORECASE,
)
ACCURACY_RE = re.compile(r"(±\s*\d+(?:[.,]\d+)?\s*%|±\s*\d+(?:[.,]\d+)?\s*[A-Za-z°]+)", re.IGNORECASE)
MATERIAL_RE = re.compile(r"(AISI\s*304|AISI\s*316L|316L|304|Hastelloy|PTFE|PVDF|PFA|SS316L|SS304)", re.IGNORECASE)
CONNECTION_RE = re.compile(r"(G\s*1/2|G1/2|NPT\s*1/2|NPT\s*1/4|DN\s*\d+\b|PN\s*\d+\b|M\d+\b)", re.IGNORECASE)
TEMP_TYPE_RE = re.compile(r"(Pt\s*1000?\b|Type\s*[KJTN]|Thermocouple\s*[KJTN])", re.IGNORECASE)
QTY_MULT_RE = re.compile(r"(?:^|\b)(?:x|×|х)\s*(\d+)\b", re.IGNORECASE)
QTY_WORD_RE = re.compile(r"\b(один|одна|две|два|три|четыре|пять|six|seven|eight|nine|ten)\b", re.IGNORECASE)

RU_NUM_WORDS = {
    "один": 1, "одна": 1, "две": 2, "два": 2, "три": 3, "четыре": 4, "пять": 5,
}

def parse_qty(text: str) -> Tuple[float, Optional[str]]:
    qty = 1.0
    uom: Optional[str] = None
    m = QTY_MULT_RE.search(text)
    if m:
        qty = safe_float(m.group(1), 1.0)
    else:
        m2 = QTY_WORD_RE.search(text)
        if m2:
            qty = float(RU_NUM_WORDS.get(m2.group(1).lower(), 1))
    # crude UoM detection (pieces)
    if re.search(r"\b(шт|pcs|pieces?)\b", text, re.IGNORECASE):
        uom = "count"
    return qty, uom

def parse_output_signals(text: Optional[str]) -> List[str]:
    if not text:
        return []
    t = text.lower()
    out = set()
    if re.search(r'\b4\s*[-–—]?\s*20\s*m?a\b', t):
        out.add("4–20 mA")
    if re.search(r'\b0\s*[-–—]?\s*10\s*v\b', t):
        out.add("0–10 V")
    if "rs-485" in t or "rs485" in t:
        out.add("RS-485")
    if "modbus" in t:
        out.add("Modbus")
    if "hart" in t:
        out.add("HART")
    if "canopen" in t:
        out.add("CANopen")
    if "profibus" in t:
        out.add("Profibus")
    if re.search(r'\bnpn\b', t):
        out.add("NPN")
    if re.search(r'\bpnp\b', t):
        out.add("PNP")
    if "relay" in t:
        out.add("Relay")
    if "pulse" in t or "frequency" in t:
        out.add("Pulse")
    return sorted(out)

def parse_ip(text: Optional[str]) -> Optional[str]:
    if not text:
        return None
    m = re.search(r'\bIP\s*(\d\d)\b', text, re.IGNORECASE)
    if m:
        return f"IP{m.group(1)}"
    return None

def parse_thread(text: Optional[str]) -> Optional[str]:
    if not text:
        return None
    m = re.search(r'\b([GM]\d/\d+|NPT\d/\d+|DN\d+)\b', text, re.IGNORECASE)
    if m:
        return m.group(1)
    return None

def split_row_into_items(
    row: pd.Series,
    idcol: str,
    textcol: str,
    keywords: Dict[str, List[str]],
    sheet: str,
    excel_path: Path,
) -> List[ItemCandidate]:
    raw_text = " ".join([str(row.get(textcol) or ""), str(row.get(idcol) or "")]).strip()
    raw_text = re.sub(r"\s+", " ", raw_text)
    line_id_raw = str(row.get(idcol) or "").strip()
    base_line_id = line_id_raw if line_id_raw else sha1_hex(f"{sheet}:{row.get('__row_index__') or ''}:{raw_text}")[:16]
    row_hash = sha1_hex(f"{sheet}:{excel_path}:{row.get('__row_index__')}:{raw_text}")

    flags: List[str] = []
    if any(re.search(rf"\b{re.escape(w)}\b", raw_text, re.IGNORECASE) for w in keywords.get("sensor_like", [])):
        flags.append("Sensor")

    is_assembly = any(w in raw_text.lower() for w in [*keywords.get("assembly", []), "в сборе", "комплект"])
    if is_assembly:
        flags.append("Assembly")

    qty, uom = parse_qty(raw_text)

    # Component splitting heuristic: split by separators after assembly markers
    components: List[str] = []
    if is_assembly:
        # try splitting by common delimiters ";", ",", "+"
        parts = re.split(r"[;,/＋\+]| и ", raw_text, flags=re.IGNORECASE)
        # Keep components that look like separate tokens (avoid over-splitting)
        components = [p.strip() for p in parts if len(p.strip()) > 5 and p.strip() != raw_text]

    items: List[ItemCandidate] = []
    parent_line_id = base_line_id
    items.append(
        ItemCandidate(
            lineId=parent_line_id,
            groupId=base_line_id,
            parentLineId=None,
            role="parent",
            qty=qty,
            uom=uom,
            specCore=raw_text,
            rawText=raw_text,
            rowHash=row_hash,
            rowIndex=int(row.get("__row_index__") or -1),
            sheet=sheet,
            excelPath=str(excel_path),
            flags=flags,
        )
    )

    for idx, comp in enumerate(components[:8], start=1):  # limit to avoid explosion
        cid = base_line_id + f".c{idx}"
        cqty, cuom = parse_qty(comp)
        items.append(
            ItemCandidate(
                lineId=cid,
                groupId=base_line_id,
                parentLineId=parent_line_id,
                role="component",
                qty=cqty,
                uom=cuom,
                specCore=comp,
                rawText=raw_text,
                rowHash=row_hash,
                rowIndex=int(row.get("__row_index__") or -1),
                sheet=sheet,
                excelPath=str(excel_path),
                flags=[f for f in flags if f != "Assembly"],
            )
        )
    return items

# --- LLM classification (Qwen via utils.load_model) ---------------------------------------------

def _call_llm_text(llm: Any, prompt: str) -> str:
    """
    Try multiple common call patterns against the loaded model.
    """
    for attr in ("generate", "invoke", "chat", "__call__"):
        if hasattr(llm, attr):
            try:
                fn: Callable = getattr(llm, attr)  # type: ignore
                out = fn(prompt)
                # Extract text for common wrapper types
                if isinstance(out, str):
                    return out
                # common patterns: {"text": "..."} or list[{"generated_text": "..."}]
                if hasattr(out, "text"):
                    return str(out.text)
                if isinstance(out, dict):
                    if "text" in out:
                        return str(out["text"])
                    if "content" in out:
                        return str(out["content"])
                if isinstance(out, list) and out:
                    cand = out[0]
                    if isinstance(cand, dict):
                        return str(cand.get("text") or cand.get("generated_text") or cand)
                return str(out)
            except Exception as e:
                logger.debug("LLM call pattern failed", extra={"extra": {"attr": attr, "error": str(e)}})
                continue
    raise RuntimeError("Unsupported LLM interface; cannot obtain text.")

def extract_first_json(s: str) -> Optional[Dict[str, Any]]:
    """
    Robustly extract the first JSON object from a string.
    """
    # Try a fast-path: find first '{' and last '}' and parse
    start = s.find("{")
    end = s.rfind("}")
    if start != -1 and end != -1 and end > start:
        fragment = s[start : end + 1]
        try:
            return orjson.loads(fragment)
        except Exception:
            pass
    # Fallback: balance braces
    depth = 0
    buf = []
    started = False
    for ch in s:
        if ch == "{":
            depth += 1
            started = True
        if started:
            buf.append(ch)
        if ch == "}":
            depth -= 1
            if started and depth == 0:
                frag = "".join(buf)
                try:
                    return orjson.loads(frag)
                except Exception:
                    return None
    return None

def llm_classify_item(
    llm: Any,
    spec_core: str,
    lang_list: List[str],
    class_dict: Dict[str, Dict[str, Any]],
    prop_dict: Dict[str, List[Dict[str, Any]]],
    eclass_version: str,
    max_retries: int = 2,
) -> ClassifiedItem:
    system_prompt = (
        "You are an expert in ECLASS Advanced (version "
        + eclass_version
        + ") classification for industrial items. "
        "Return STRICT JSON with fields: "
        '{"eclass_code":"XX-XX-XX-XX","eclass_label":"...",'
        '"confidence":0.0,"properties":[{"irdi":"...","name":"...","value":"...",'
        '"unit":"...","datatype":"number|string|boolean|enum"}],"notes":"..."}.\n'
        "Rules:\n"
        "- Choose the most specific 8-digit class code (format XX-XX-XX-XX).\n"
        "- Only include properties defined for that class in the provided dictionary (if known to you).\n"
        "- Use SI units where appropriate.\n"
        "- Always respond with JSON even if unsure (set confidence accordingly)."
    )
    classes_hint = ", ".join(list(itertools.islice(class_dict.keys(), 20)))
    user_prompt = (
        f"[LANGS]={','.join(lang_list)}\n"
        f"[SPEC]: {spec_core}\n"
        f"[KNOWN_CLASS_CODES_SAMPLE]: {classes_hint}\n"
        "Respond with STRICT JSON only."
    )
    prompt = system_prompt + "\n\n" + user_prompt

    # Default fallback result (low confidence)
    fallback = ClassifiedItem(
        lineId="",
        eclass_code=None,
        eclass_label=None,
        confidence=0.0,
        properties=[],
        notes="fallback",
    )

    if llm is None:
        # utils.load_model missing; fallback path will run later
        return fallback

    # Attempt calls with a couple retries
    for attempt in range(max_retries + 1):
        try:
            raw = _call_llm_text(llm, prompt)
            logger.debug("LLM raw response", extra={"extra": {"resp": raw[:800]}})
            data = extract_first_json(raw)
            if not data or not isinstance(data, dict):
                raise ValueError("LLM did not return JSON.")

            eclass_code = data.get("eclass_code")
            eclass_label = data.get("eclass_label")
            conf = float(data.get("confidence") or 0.0)
            props_in = data.get("properties") or []
            props: List[PropertyValueCandidate] = []

            # Filter properties to those defined for that class (if available)
            allowed = {p["irdi"]: p for p in prop_dict.get(str(eclass_code or ""), [])}
            allowed_by_name = {p["name"].lower(): p for p in prop_dict.get(str(eclass_code or ""), [])}

            for p in props_in:
                try:
                    irdi = str(p.get("irdi") or "").strip() or None
                    name = str(p.get("name") or "").strip()
                    datatype = str(p.get("datatype") or "string").strip()
                    value = p.get("value")
                    unit = p.get("unit") or None

                    pdict = None
                    if irdi and irdi in allowed:
                        pdict = allowed[irdi]
                    elif name and name.lower() in allowed_by_name:
                        pdict = allowed_by_name[name.lower()]

                    props.append(
                        PropertyValueCandidate(
                            irdi=pdict["irdi"] if pdict else irdi,
                            name=pdict["name"] if pdict else name,
                            value=value,
                            unit=unit,
                            datatype=datatype if datatype in {"number", "string", "boolean", "enum"} else "string",
                            source="llm",
                            confidence=conf,
                            canonicalUnit=(pdict.get("canonicalUnit") if pdict else None) if isinstance(pdict, dict) else None,
                        )
                    )
                except Exception as e:
                    logger.debug("Skipping malformed property from LLM", extra={"extra": {"error": str(e)}})

            return ClassifiedItem(
                lineId="",  # to be filled by caller
                eclass_code=eclass_code if eclass_code in class_dict else None,
                eclass_label=eclass_label,
                confidence=conf,
                properties=props,
                notes=str(data.get("notes") or ""),
            )
        except Exception as e:
            if attempt < max_retries:
                time.sleep(0.5 * (2 ** attempt))
                continue
            logger.debug("LLM classification failed; will fallback", extra={"extra": {"error": str(e)}})
            return fallback

# --- Regex heuristics for properties ------------------------------------------------------------

def heuristic_properties(
    spec_core: str,
    class_code: Optional[str],
    prop_dict: Dict[str, List[Dict[str, Any]]],
) -> List[PropertyValueCandidate]:
    """
    Extract a few common properties using regex and map to class property model by fuzzy name match.
    """
    candidates: List[Tuple[str, Any, Optional[str], str]] = []  # (name, value, unit, datatype)

    m = RANGE_RE.search(spec_core)
    if m:
        minv = m.group("min")
        maxv = m.group("max")
        unit = (m.group("unit") or "").strip() or None
        # Represent as a text range if we don't have a specific prop; caller can map
        candidates.append(("Measuring range", f"{minv}..{maxv}", unit, "string"))

    sigs = OUTPUT_SIG_RE.findall(spec_core)
    if sigs:
        candidates.append(("Output signal", " / ".join({re.sub(r"\s+", "", s) for s in sigs}), None, "enum"))

    acc = ACCURACY_RE.search(spec_core)
    if acc:
        candidates.append(("Accuracy", acc.group(1), None, "string"))

    mats = MATERIAL_RE.findall(spec_core)
    if mats:
        candidates.append(("Material", " / ".join(sorted(set(mats))), None, "enum"))

    conn = CONNECTION_RE.findall(spec_core)
    if conn:
        candidates.append(("Process connection", " / ".join(sorted(set(conn))), None, "enum"))

    ttype = TEMP_TYPE_RE.search(spec_core)
    if ttype:
        candidates.append(("Temperature sensor type", ttype.group(1), None, "enum"))

    # Map names to class properties via fuzzy matching (if available)
    out: List[PropertyValueCandidate] = []
    if class_code and class_code in prop_dict:
        model_names = [p["name"] for p in prop_dict[class_code]]
        for name, value, unit, dtype in candidates:
            match = rf_process.extractOne(name, model_names, scorer=fuzz.WRatio)
            pdict = None
            if match and match[1] >= 80:
                mname = match[0]
                pdict = next((p for p in prop_dict[class_code] if p["name"] == mname), None)
            out.append(
                PropertyValueCandidate(
                    irdi=pdict["irdi"] if pdict else None,
                    name=pdict["name"] if pdict else name,
                    value=value,
                    unit=unit,
                    datatype=dtype,
                    source="regex",
                    confidence=0.4,
                    canonicalUnit=pdict.get("canonicalUnit") if pdict else None,
                )
            )
    else:
        for name, value, unit, dtype in candidates:
            out.append(
                PropertyValueCandidate(
                    irdi=None, name=name, value=value, unit=unit, datatype=dtype, source="regex", confidence=0.3
                )
            )
    return out

# --- Unit normalization -------------------------------------------------------------------------

def normalize_units(
    prop: PropertyValueCandidate,
    ureg: UnitRegistry,
    unit_aliases: Dict[str, str],
) -> PropertyValueCandidate:
    """
    Convert prop.value & unit to canonical units when datatype is number or numeric-looking and a canonicalUnit exists.
    Stores canonicalUnit in the returned candidate; keeps original unit in 'unit'.
    """
    # Try to coerce numeric
    def _to_float(val: Any) -> Optional[float]:
        if isinstance(val, (int, float)):
            return float(val)
        if isinstance(val, str):
            try:
                return float(val.replace(",", "."))
            except Exception:
                # detect range "a..b" -> average
                mm = re.match(r"\s*(-?\d+(?:[.,]\d+)?)\s*\.\.\s*(-?\d+(?:[.,]\d+)?)", val)
                if mm:
                    a = float(mm.group(1).replace(",", "."))
                    b = float(mm.group(2).replace(",", "."))
                    return (a + b) / 2.0
        return None

    if prop.canonicalUnit:
        # normalize unit symbols
        unit_in = prop.unit
        if isinstance(unit_in, str) and unit_in in unit_aliases:
            unit_in = unit_aliases[unit_in]

        num = _to_float(prop.value)
        if num is not None and isinstance(prop.canonicalUnit, str):
            try:
                if unit_in:
                    q = num * ureg(unit_in)
                else:
                    # If unspecified unit and canonical is dimensionless or same symbol, keep as-is
                    q = num * ureg(prop.canonicalUnit)
                q_to = q.to(prop.canonicalUnit)
                return PropertyValueCandidate(
                    **{**asdict(prop), "value": float(q_to.magnitude), "unit": prop.canonicalUnit}
                )
            except Exception:
                # leave as-is if conversion failed
                return prop
    return prop

# --- Neo4j: connection, constraints, and batched upserts ----------------------------------------

CONSTRAINT_QUERIES = [
    "CREATE CONSTRAINT eclass_class_unique IF NOT EXISTS FOR (n:ECLASSClass) REQUIRE (n.code, n.version) IS UNIQUE",
    "CREATE CONSTRAINT property_unique IF NOT EXISTS FOR (n:Property) REQUIRE (n.irdi, n.version) IS UNIQUE",
    "CREATE CONSTRAINT unit_unique IF NOT EXISTS FOR (n:Unit) REQUIRE (n.symbol) IS UNIQUE",
    "CREATE CONSTRAINT item_unique IF NOT EXISTS FOR (n:Item) REQUIRE (n.lineId) IS UNIQUE",
    "CREATE CONSTRAINT sourcerow_unique IF NOT EXISTS FOR (n:SourceRow) REQUIRE (n.hash) IS UNIQUE",
]

def connect_neo4j_from_env() -> Tuple[Optional[Driver], Optional[str]]:
    """
    Connect to Neo4j using environment variables from .env file.
    Returns (driver, database) tuple or (None, None) on failure.
    """
    # Explicitly load from .env file in the project root
    load_dotenv(Path(__file__).parent / ".env")
    
    try:
        # Get connection details from environment variables
        uri = os.environ.get("NEO4J_URI")
        user = os.environ.get("NEO4J_USER")
        pwd = os.environ.get("NEO4J_PASSWORD")
        database = os.environ.get("NEO4J_DATABASE")
        
        if not (uri and user and pwd):
            logger.error("Missing Neo4j connection details in .env file")
            return None, None
        
        driver = GraphDatabase.driver(uri, auth=(user, pwd))
        # Verify connectivity
        driver.verify_connectivity()
        
        with driver.session(database=database) as sess:
            ok = sess.run("RETURN 1 AS ok").single()
            if not ok or ok["ok"] != 1:
                logger.warning("Neo4j health check failed")
                return None, None
        
        logger.info("Connected to Neo4j", extra={
            "uri": uri.replace(f"{user}:{pwd}", f"{user}:***"), 
            "database": database
        })
        return driver, database
    except Exception as e:
        logger.error(f"Failed to connect to Neo4j: {str(e)}")
        return None, None

def ensure_constraints(driver: Optional[Driver], database: Optional[str]) -> None:
    if not driver or not database:
        logger.warning("Cannot ensure constraints: no driver/database connection")
        return
        
    with driver.session(database=database) as sess:
        for q in CONSTRAINT_QUERIES:
            sess.run(q)

def upsert_eclass_dictionaries(
    driver: Optional[Driver],
    database: Optional[str],
    classes: Dict[str, Dict[str, Any]],
    props_by_class: Dict[str, List[Dict[str, Any]]],
) -> None:
    if not driver or not database:
        logger.warning("Cannot upsert dictionaries: no driver/database connection")
        return
    
    # Upsert classes
    if classes:
        payload = [{"code": c["code"], "label": c.get("label", ""), "version": c.get("version", "15.0")} for c in classes.values()]
        q = """
        UNWIND $rows AS r
        MERGE (c:ECLASSClass {code: r.code, version: r.version})
        SET c.label = r.label
        """
        with driver.session(database=database) as sess:
            sess.run(q, rows=payload)

    # Upsert properties & HAS_PROPERTY_MODEL
    rows: List[Dict[str, Any]] = []
    for class_code, plist in props_by_class.items():
        for p in plist:
            rows.append({
                "class_code": class_code,
                "version": p.get("version", "15.0"),
                "irdi": p.get("irdi"),
                "name": p.get("name"),
                "datatype": p.get("datatype"),
                "canonicalUnit": p.get("canonicalUnit"),
            })
    if rows:
        q = """
        UNWIND $rows AS r
        MERGE (c:ECLASSClass {code: r.class_code, version: r.version})
        MERGE (p:Property {irdi: r.irdi, version: r.version})
        SET p.name = r.name, p.datatype = r.datatype, p.canonicalUnit = r.canonicalUnit
        MERGE (c)-[:HAS_PROPERTY_MODEL]->(p)
        WITH DISTINCT r.canonicalUnit AS sym
        WHERE sym IS NOT NULL
        MERGE (:Unit {symbol: sym})
        """
        with driver.session(database=database) as sess:
            sess.run(q, rows=rows)

def upsert_items_batch(
    driver: Driver,
    database: str,
    items: List[ItemCandidate],
    class_map: Dict[str, Dict[str, Any]],
    classified: Dict[str, ClassifiedItem],
    prop_dict: Dict[str, List[Dict[str, Any]]],
    confidence_threshold: float,
) -> Dict[str, int]:
    """
    Write Items, SourceRows, relationships and PropertyValue nodes in batches.
    """
    # Prepare rows for Items & SourceRows
    item_rows: List[Dict[str, Any]] = []
    sr_rows: Dict[str, Dict[str, Any]] = {}
    instance_rows: List[Dict[str, Any]] = []
    review_links: List[Dict[str, Any]] = []
    component_links: List[Dict[str, Any]] = []
    accessory_links: List[Dict[str, Any]] = []
    service_links: List[Dict[str, Any]] = []
    propval_rows: List[Dict[str, Any]] = []
    unit_rows: Dict[str, Dict[str, Any]] = {}

    # Build a (class_code -> version) quick map
    class_version_map = {k: v.get("version", "15.0") for k, v in class_map.items()}

    for it in items:
        ci = classified.get(it.lineId)
        # SourceRow
        sr_rows[it.rowHash] = {
            "hash": it.rowHash,
            "rowId": it.rowIndex,
            "sheet": it.sheet,
            "excelPath": it.excelPath,
            "rawText": it.rawText,
        }
        # Item
        item_row = {
            "lineId": it.lineId,
            "role": it.role,
            "specCore": it.specCore,
            "qty": it.qty,
            "uom": it.uom,
            "flags": list(sorted(set(it.flags))),
            "rowHash": it.rowHash,
        }
        item_rows.append(item_row)

        # Instance_of relationship
        if ci and ci.eclass_code:
            instance_rows.append({
                "lineId": it.lineId,
                "class_code": ci.eclass_code,
                "class_version": class_version_map.get(ci.eclass_code, "15.0"),
            })

        # Review link condition
        low_conf = (not ci) or (not ci.eclass_code) or (ci.confidence < confidence_threshold)
        if low_conf:
            review_links.append({"lineId": it.lineId, "hash": it.rowHash})

        # Parent-child relations
        if it.parentLineId and it.role == "component":
            component_links.append({"parentId": it.parentLineId, "childId": it.lineId, "qty": it.qty})
        elif it.parentLineId and it.role == "accessory":
            accessory_links.append({"mainId": it.parentLineId, "accId": it.lineId})
        elif it.parentLineId and it.role == "service":
            service_links.append({"mainId": it.parentLineId, "svcId": it.lineId})

        # Property values
        if ci:
            for pv in ci.properties:
                # Prepare unit nodes
                if pv.unit:
                    unit_rows[pv.unit] = {"symbol": pv.unit}

                # Determine Property node version
                pver = None
                if ci.eclass_code and ci.eclass_code in prop_dict and pv.irdi:
                    # Look up property version in dict
                    for p in prop_dict[ci.eclass_code]:
                        if p.get("irdi") == pv.irdi:
                            pver = p.get("version", "15.0")
                            break
                if pver is None:
                    pver = "15.0"

                # Prepare value fields and pvId
                num_val = None
                text_val = None
                bool_val = None
                enum_val = None
                if pv.datatype == "number" and isinstance(pv.value, (int, float)):
                    num_val = float(pv.value)
                elif pv.datatype == "boolean":
                    bool_val = bool(pv.value)
                elif pv.datatype == "enum":
                    enum_val = str(pv.value)
                else:
                    text_val = str(pv.value)

                pv_id = sha1_hex(f"{it.lineId}|{pv.irdi or pv.name}|{num_val or text_val or bool_val or enum_val}|{pv.unit}|{pv.source}")

                propval_rows.append({
                    "pvId": pv_id,
                    "lineId": it.lineId,
                    "irdi": pv.irdi,
                    "pversion": pver,
                    "pname": pv.name,
                    "numericValue": num_val,
                    "textValue": text_val,
                    "boolValue": bool_val,
                    "enumValue": enum_val,
                    "unit": pv.unit,
                    "source": pv.source,
                    "confidence": float(pv.confidence),
                })

    database_param = database  # shadow for closure

    with driver.session(database=database_param) as sess:
        # SourceRows
        if sr_rows:
            q_sr = """
            UNWIND $rows AS r
            MERGE (s:SourceRow {hash: r.hash})
            SET s.rowId = r.rowId, s.sheet = r.sheet, s.excelPath = r.excelPath, s.rawText = r.rawText
            """
            sess.run(q_sr, rows=list(sr_rows.values()))

        # Items & DERIVED_FROM
        if item_rows:
            q_item = """
            UNWIND $rows AS r
            MERGE (i:Item {lineId: r.lineId})
              ON CREATE SET i.role = r.role, i.specCore = r.specCore, i.qty = r.qty, i.uom = r.uom, i.flags = r.flags
              ON MATCH  SET i.role = r.role, i.specCore = r.specCore, i.qty = r.qty, i.uom = r.uom, i.flags = r.flags
            WITH r
            MATCH (s:SourceRow {hash: r.rowHash})
            MATCH (i:Item {lineId: r.lineId})
            MERGE (i)-[:DERIVED_FROM]->(s)
            """
            sess.run(q_item, rows=item_rows)

        # INSTANCE_OF
        if instance_rows:
            q_inst = """
            UNWIND $rows AS r
            MATCH (i:Item {lineId: r.lineId})
            MERGE (c:ECLASSClass {code: r.class_code, version: r.class_version})
            MERGE (i)-[:INSTANCE_OF]->(c)
            """
            sess.run(q_inst, rows=instance_rows)

        # Relations: components/accessories/services
        if component_links:
            q_comp = """
            UNWIND $rows AS r
            MATCH (p:Item {lineId: r.parentId})
            MATCH (c:Item {lineId: r.childId})
            MERGE (p)-[rel:HAS_COMPONENT]->(c)
            SET rel.qty = r.qty
            """
            sess.run(q_comp, rows=component_links)
        if accessory_links:
            q_acc = """
            UNWIND $rows AS r
            MATCH (a:Item {lineId: r.accId})
            MATCH (m:Item {lineId: r.mainId})
            MERGE (a)-[:ACCESSORY_OF]->(m)
            """
            sess.run(q_acc, rows=accessory_links)
        if service_links:
            q_svc = """
            UNWIND $rows AS r
            MATCH (svc:Item {lineId: r.svcId})
            MATCH (m:Item {lineId: r.mainId})
            MERGE (svc)-[:SERVICE_FOR]->(m)
            """
            sess.run(q_svc, rows=service_links)

        # Units
        if unit_rows:
            q_unit = """
            UNWIND $rows AS r
            MERGE (:Unit {symbol: r.symbol})
            """
            sess.run(q_unit, rows=list(unit_rows.values()))

        # PropertyValues
        if propval_rows:
            q_pv = """
            UNWIND $rows AS r
            MATCH (i:Item {lineId: r.lineId})
            OPTIONAL MATCH (p:Property {irdi: r.irdi, version: r.pversion})
            MERGE (v:PropertyValue {pvId: r.pvId})
              ON CREATE SET v.numericValue = r.numericValue, v.textValue = r.textValue, v.boolValue = r.boolValue,
                            v.enumValue = r.enumValue, v.unit = r.unit, v.source = r.source, v.confidence = r.confidence
              ON MATCH  SET v.numericValue = r.numericValue, v.textValue = r.textValue, v.boolValue = r.boolValue,
                            v.enumValue = r.enumValue, v.unit = r.unit, v.source = r.source, v.confidence = r.confidence
            MERGE (i)-[:HAS_PROPERTY]->(v)
            FOREACH(_ IN CASE WHEN p IS NULL THEN [] ELSE [1] END |
              MERGE (v)-[:OF_PROPERTY]->(p)
            )
            WITH r, v
            FOREACH(_ IN CASE WHEN r.unit IS NULL THEN [] ELSE [1] END |
              MERGE (u:Unit {symbol: r.unit})
              MERGE (v)-[:USES_UNIT]->(u)
            )
            """
            sess.run(q_pv, rows=propval_rows)

        # Review links
        if review_links:
            q_rev = """
            UNWIND $rows AS r
            MATCH (i:Item {lineId: r.lineId})
            MATCH (s:SourceRow {hash: r.hash})
            MERGE (i)-[:REQUIRES_REVIEW]->(s)
            """
            sess.run(q_rev, rows=review_links)

    return {
        "items": len(item_rows),
        "props": len(propval_rows),
        "reviews": len(review_links),
        "components": len(component_links),
        "accessories": len(accessory_links),
        "services": len(service_links),
    }

# --- Pipeline -----------------------------------------------------------------------------------

def process_excel(
    excel_path: Path,
    sheet: str,
    idcol: str,
    classes: Dict[str, Dict[str, Any]],
    props_by_class: Dict[str, List[Dict[str, Any]]],
    confidence_threshold: float,
    batch_size: int,
    langs: List[str],
    dry_run: bool,
    cfg: Dict[str, Any],
    ureg: UnitRegistry,
    database: Optional[str] = None,
    driver: Optional[Driver] = None,
) -> Dict[str, int]:
    """
    Main ETL pipeline over an Excel sheet.
    """
    # LLM model (optional)
    llm = None
    if load_model:
        try:
            llm = load_model("qwen2.5-14b-instruct-4bit")
        except Exception as e:
            logger.warning(
                "Failed to load LLM via utils.load_model; fallback to regex only",
                extra={"extra": {"error": str(e)}},
            )
    else:
        if _UTILS_IMPORT_ERROR:
            logger.warning(
                "utils.load_model not available; fallback to regex only",
                extra={"extra": {"error": str(_UTILS_IMPORT_ERROR)}})
    unit_aliases = cfg.get("unit_aliases", {})
    eclass_version = str(cfg.get("eclass_version", "15.0"))
    columns = cfg.get("columns", {})
    text_candidates = columns.get("text_candidates", [])

    items_out: List[Dict[str, Any]] = []
    props_out: List[Dict[str, Any]] = []

    summaries = {
        "rows_processed": 0,
        "items_upserted": 0,
        "props_upserted": 0,
        "review_count": 0,
        "components": 0,
        "accessories": 0,
        "services": 0,
    }

    chunk_iter = read_excel_chunks(excel_path, sheet, chunk_size=batch_size)
    for df in tqdm(chunk_iter, desc="Reading Excel in chunks"):
        # Determine columns
        textcol = guess_text_column(df, text_candidates)
        if not textcol:
            # fallback to first non-id column
            non_id_cols = [c for c in df.columns if c != idcol and not str(c).startswith("__")]
            textcol = non_id_cols[0] if non_id_cols else idcol
        if idcol not in df.columns:
            df[idcol] = None

        # Process rows -> ItemCandidates
        chunk_items: List[ItemCandidate] = []
        for _, row in df.iterrows():
            if row.isna().all():
                continue
            chunk_items.extend(
                split_row_into_items(
                    row=row,
                    idcol=idcol,
                    textcol=textcol,
                    keywords=cfg.get("keywords", {}),
                    sheet=sheet,
                    excel_path=excel_path,
                )
            )

        summaries["rows_processed"] += len(df)

        # Classify each unique lineId (parent and components)
        classified_by_id: Dict[str, ClassifiedItem] = {}
        for it in chunk_items:
            # Prepare classification
            ci = llm_classify_item(
                llm=llm,
                spec_core=it.specCore,
                lang_list=langs,
                class_dict=classes,
                prop_dict=props_by_class,
                eclass_version=eclass_version,
            )
            ci.lineId = it.lineId

            # Heuristic properties & merging
            if not ci.eclass_code:
                # If model didn't produce a valid class, attempt to infer the closest by keyword match
                # (kept minimal; still mark low confidence)
                ci.eclass_code = None
                ci.confidence = min(ci.confidence, 0.3)
            # Merge regex heuristics props (avoid duplicates by name/irdi)
            heur = heuristic_properties(it.specCore, ci.eclass_code, props_by_class)
            existing_keys = {(p.irdi, p.name.lower()) for p in ci.properties}
            for hp in heur:
                key = (hp.irdi, hp.name.lower())
                if key not in existing_keys:
                    ci.properties.append(hp)
                    existing_keys.add(key)

            # Normalize units to canonical where possible
            ci_norm_props: List[PropertyValueCandidate] = []
            for p in ci.properties:
                ci_norm_props.append(normalize_units(p, ureg, unit_aliases))
            ci.properties = ci_norm_props

            classified_by_id[it.lineId] = ci

        # Write batch to DB (unless dry-run)
        if not dry_run:
            if not (driver and database):
                raise RuntimeError("Driver/database required for DB writes.")
            stats = upsert_items_batch(
                driver=driver,
                database=database,
                items=chunk_items,
                class_map=classes,
                classified=classified_by_id,
                prop_dict=props_by_class,
                confidence_threshold=confidence_threshold,
            )
            summaries["items_upserted"] += stats["items"]
            summaries["props_upserted"] += stats["props"]
            summaries["review_count"] += stats["reviews"]
            summaries["components"] += stats["components"]
            summaries["accessories"] += stats["accessories"]
            summaries["services"] += stats["services"]

        # Collect outputs for dry-run artifacts (and for audit even when writing)
        for it in chunk_items:
            ci = classified_by_id.get(it.lineId)
            items_out.append(
                {
                    "lineId": it.lineId,
                    "role": it.role,
                    "qty": it.qty,
                    "uom": it.uom,
                    "specCore": it.specCore,
                    "flags": "|".join(sorted(set(it.flags))),
                    "eclass_code": ci.eclass_code if ci else None,
                    "eclass_label": ci.eclass_label if ci else None,
                    "confidence": ci.confidence if ci else 0.0,
                    "notes": ci.notes if ci else "",
                    "rowHash": it.rowHash,
                    "rowIndex": it.rowIndex,
                    "sheet": it.sheet,
                    "excelPath": it.excelPath,
                }
            )
            if ci:
                for p in ci.properties:
                    props_out.append(
                        {
                            "lineId": it.lineId,
                            "irdi": p.irdi,
                            "name": p.name,
                            "value": p.value,
                            "unit": p.unit,
                            "datatype": p.datatype,
                            "source": p.source,
                            "confidence": p.confidence,
                            "canonicalUnit": p.canonicalUnit,
                        }
                    )

    # Dry-run output files
    out_base = excel_path.with_suffix("")
    items_df = pd.DataFrame(items_out)
    props_df = pd.DataFrame(props_out)
    items_csv = out_base.parent / f"{out_base.name}_items.csv"
    props_csv = out_base.parent / f"{out_base.name}_props.csv"
    items_df.to_csv(items_csv, index=False)
    props_df.to_csv(props_csv, index=False)

    # Try Parquet if available (optional; do not enforce pyarrow)
    try:
        items_parq = out_base.parent / f"{out_base.name}_items.parquet"
        props_parq = out_base.parent / f"{out_base.name}_props.parquet"
        items_df.to_parquet(items_parq, index=False)
        props_df.to_parquet(props_parq, index=False)
    except Exception as e:
        logger.debug("Parquet export skipped", extra={"extra": {"reason": str(e)}})

    logger.info(
        "Batch summary",
        extra={
            "extra": {
                **summaries,
                "dry_run_outputs": {"items_csv": str(items_csv), "props_csv": str(props_csv)},
            }
        },
    )
    return summaries

# --- CLI ----------------------------------------------------------------------------------------

def build_argparser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="ECLASS Neo4j graph loader from Excel")
    p.add_argument("--excel", type=Path, help="Path to Excel workbook (.xlsx)")
    p.add_argument("--sheet", type=str, default="ПКШ-ОБОР", help="Sheet name (default: ПКШ-ОБОР)")
    p.add_argument("--idcol", type=str, default="Идентификатор", help="ID column name (default: Идентификатор)")
    p.add_argument("--config", type=Path, help="YAML/JSON config for column mapping & keyword dictionaries")
    p.add_argument("--eclass-classes", type=Path, dest="classes", help="ECLASS classes (CSV/JSON)")
    p.add_argument("--eclass-props", type=Path, dest="props", help="ECLASS properties by class (CSV/JSON)")
    p.add_argument("--batch-size", type=int, default=1000, help="Chunk size for reading/writing (default: 1000)")
    p.add_argument("--lang", type=str, default="ru,en", help="Comma-separated languages (default: ru,en)")
    p.add_argument("--confidence-threshold", type=float, default=0.75, help="LLM confidence threshold")
    p.add_argument("--dry-run", action="store_true", help="Do not write to DB; export CSV/Parquet only")
    p.add_argument("--check-db", action="store_true", help="Check DB connectivity & constraints then exit")
    p.add_argument("-v", "--verbose", action="count", default=0, help="Increase logging verbosity")
    return p

def main() -> None:
    args = build_argparser().parse_args()
    configure_logging(args.verbose)

    # Check DB only
    if args.check_db:
        driver, database = connect_neo4j_from_env()
        if driver is None:
            logger.error("DB check failed - could not connect to Neo4j")
            sys.exit(1) if args.check_db else None
        else:
            ensure_constraints(driver, database)
            logger.info("DB check OK", extra={"extra": {"uri": mask_secret(os.environ.get("NEO4J_URI", "")), 
                       "database": os.environ.get("NEO4J_DATABASE", "")}})
            driver.close()
            return

    # If no excel provided, run a tiny fixture under dry-run to validate the pipeline
    if not args.excel:
        args.dry_run = True
        tmp_excel = Path.cwd() / "fixture.xlsx"
        # Create an in-memory DataFrame and save
        df = pd.DataFrame(
            {
                "Идентификатор": ["L1", "L2", None],
                "Описание": [
                    "Датчик давления 0–10 бар, выход 4–20 мА, AISI 316L, G1/2, точность ±0.5 %",
                    "Thermocouple Type K, range 0–400 °C, sheath SS304, NPT 1/2",
                    "Комплект: датчик уровня + кабель 10 м, 4–20mA, Hastelloy",
                ],
            }
        )
        with pd.ExcelWriter(tmp_excel, engine="openpyxl") as w:
            df.to_excel(w, sheet_name="ПКШ-ОБОР", index=False)
        args.excel = tmp_excel
        logger.info("No --excel provided; using inline fixture in dry-run mode", extra={"extra": {"excel": str(tmp_excel)}})

    cfg = load_config(args.config)
    classes = load_eclass_classes(args.classes)
    props_by_class = load_eclass_properties(args.props)

    ureg = UnitRegistry(auto_reduce_dimensions=True)
    # Add a count unit (dimensionless) for pieces
    try:
        ureg.define("count = [] = count")
    except Exception:
        pass

    langs = to_list_langs(args.lang)

    driver: Optional[Driver] = None
    database: Optional[str] = None
    if not args.dry_run:
        driver, database = connect_neo4j_from_env()
        if driver is None:
            logger.warning("Neo4j connection failed - switching to dry-run mode")
            args.dry_run = True
        else:
            ensure_constraints(driver, database)
            # Upsert dictionaries first (safe/idempotent)
            upsert_eclass_dictionaries(driver, database, classes, props_by_class)

    try:
        stats = process_excel(
            excel_path=Path(args.excel),
            sheet=args.sheet,
            idcol=args.idcol,
            classes=classes,
            props_by_class=props_by_class,
            confidence_threshold=float(args.confidence_threshold),
            batch_size=int(args.batch_size),
            langs=langs,
            dry_run=bool(args.dry_run),
            cfg=cfg,
            ureg=ureg,
            database=database,
            driver=driver,
        )
        logger.info("Processing complete", extra={"extra": stats})
    finally:
        if driver:
            driver.close()

if __name__ == "__main__":
    main()
